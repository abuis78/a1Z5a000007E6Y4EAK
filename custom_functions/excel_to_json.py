def excel_to_json(vault_id=None, container_id=None, **kwargs):
    """
    Args:
        vault_id
        container_id
    
    Returns a JSON-serializable object that implements the configured data paths:
        
    """
    ############################ Custom Code Goes Below This Line #################################
    import json
    import phantom.rules as phantom
    from openpyxl import load_workbook
    from json import dumps
    
    outputs = {}
    
    # Write your custom code here...
    success, message, info = phantom.vault_info(vault_id=vault_id, container_id=container_id)
    phantom.debug(info[0]["path"])
    phantom.debug(info[0]["name"])
    file = info[0]["path"] +".xlsx"
    phantom.debug(file)
    
    wb = load_workbook(file)
    ws = wb.active
    rows = ws.max_row
    columns = ws.max_column
    
    phantom.debug(rows, columns)

    
    
    # Return a JSON-serializable object
    assert json.dumps(outputs)  # Will raise an exception if the :outputs: object is not JSON-serializable
    return outputs
